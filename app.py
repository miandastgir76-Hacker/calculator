"""
Personal Shield Premium Calculator
-----------------------------------
Flask backend: serves the SPA, performs age calculation + premium lookup
from ratecard.json, and generates PDF / Excel quotation exports.

Python 3.14.6
"""

from __future__ import annotations

import io
import json
import os
from datetime import date, datetime

from flask import (
    Flask,
    jsonify,
    render_template,
    request,
    send_file,
    send_from_directory,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RATECARD_PATH = os.path.join(BASE_DIR, "ratecard.json")

app = Flask(__name__)

MIN_AGE = 3
MAX_AGE = 80


# ---------------------------------------------------------------------------
# Rate card loading
# ---------------------------------------------------------------------------
def load_ratecard() -> dict:
    """Load the rate card JSON from disk (never hardcoded in Python or JS)."""
    with open(RATECARD_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def calculate_ages(dob: date, as_of: date | None = None) -> tuple[int, int]:
    """Return (current_age, age_next_birthday) for a given date of birth."""
    today = as_of or date.today()
    current_age = today.year - dob.year - (
        (today.month, today.day) < (dob.month, dob.day)
    )
    age_next_birthday = current_age + 1
    return current_age, age_next_birthday


def find_premium(ratecard: dict, age_next_birthday: int, treatment_limit: str):
    """Find the premium band matching age_next_birthday, then the premium
    for the requested treatment_limit within that band."""
    for band in ratecard["premium_table"]:
        if band["age_min"] <= age_next_birthday <= band["age_max"]:
            premium = band["premiums"].get(treatment_limit)
            if premium is None:
                return None, None
            return premium, band
    return None, None


def build_quote(dob_str: str, treatment_limit: str) -> dict:
    """Validate inputs and build a full quotation dict, or raise ValueError."""
    ratecard = load_ratecard()

    if treatment_limit not in ratecard["treatment_limits"]:
        raise ValueError("Invalid treatment limit selected.")

    try:
        dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        raise ValueError("Date of birth must be in YYYY-MM-DD format.")

    today = date.today()
    if dob > today:
        raise ValueError("Date of birth cannot be in the future.")

    current_age, age_next_birthday = calculate_ages(dob, today)

    if age_next_birthday < MIN_AGE or age_next_birthday > MAX_AGE:
        raise ValueError(
            f"No premium available for age next birthday {age_next_birthday}. "
            f"Coverage is available for ages {MIN_AGE} to {MAX_AGE}."
        )

    premium, band = find_premium(ratecard, age_next_birthday, treatment_limit)
    if premium is None:
        raise ValueError("No matching premium found for the given inputs.")

    annual_limit = ratecard["treatment_limits"][treatment_limit]

    return {
        "dob": dob_str,
        "current_age": current_age,
        "age_next_birthday": age_next_birthday,
        "treatment_limit": treatment_limit,
        "annual_limit": annual_limit,
        "premium": premium,
        "age_band": f"{band['age_min']}-{band['age_max']}",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# Serve PWA files from project root (correct scope for service worker)
@app.route("/manifest.json")
def manifest():
    return send_from_directory(BASE_DIR, "manifest.json", mimetype="application/manifest+json")


@app.route("/service-worker.js")
def service_worker():
    return send_from_directory(BASE_DIR, "service-worker.js", mimetype="application/javascript")


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------
@app.route("/api/ratecard", methods=["GET"])
def api_ratecard():
    """Expose treatment limit names/annual limits to populate the frontend
    dropdown dynamically (premium values themselves are never sent upfront)."""
    ratecard = load_ratecard()
    return jsonify({"treatment_limits": ratecard["treatment_limits"]})


@app.route("/api/calculate", methods=["POST"])
def api_calculate():
    data = request.get_json(silent=True) or {}
    dob_str = data.get("dob", "")
    treatment_limit = data.get("treatment_limit", "")

    try:
        quote = build_quote(dob_str, treatment_limit)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    return jsonify({"quote": quote})


@app.route("/api/export/pdf", methods=["POST"])
def api_export_pdf():
    data = request.get_json(silent=True) or {}
    dob_str = data.get("dob", "")
    treatment_limit = data.get("treatment_limit", "")

    try:
        quote = build_quote(dob_str, treatment_limit)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=28 * mm,
        bottomMargin=20 * mm,
        leftMargin=22 * mm,
        rightMargin=22 * mm,
        title="Personal Shield Premium Quotation",
    )

    rose = colors.HexColor("#9c5a68")
    slate = colors.HexColor("#5c6b73")
    green = colors.HexColor("#3b8f5c")
    light_bg = colors.HexColor("#f4f6f7")

    title_style = ParagraphStyle(
        "TitleStyle", fontName="Helvetica-Bold", fontSize=20, textColor=rose, spaceAfter=2
    )
    subtitle_style = ParagraphStyle(
        "SubtitleStyle", fontName="Helvetica", fontSize=10.5, textColor=slate, spaceAfter=18
    )
    section_style = ParagraphStyle(
        "SectionStyle", fontName="Helvetica-Bold", fontSize=12, textColor=slate, spaceBefore=10, spaceAfter=8
    )
    footer_style = ParagraphStyle(
        "FooterStyle", fontName="Helvetica-Oblique", fontSize=8.5, textColor=slate, spaceBefore=18
    )

    elements = []
    elements.append(Paragraph("Personal Shield", title_style))
    elements.append(Paragraph("Premium Quotation Summary", subtitle_style))

    elements.append(Paragraph("Applicant Details", section_style))
    detail_rows = [
        ["Date of Birth", quote["dob"]],
        ["Current Age", str(quote["current_age"])],
        ["Age Next Birthday", str(quote["age_next_birthday"])],
    ]
    detail_table = Table(detail_rows, colWidths=[70 * mm, 90 * mm])
    detail_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), slate),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#e7ebed")),
    ]))
    elements.append(detail_table)

    elements.append(Paragraph("Plan Details", section_style))
    plan_rows = [
        ["Treatment Limit", quote["treatment_limit"]],
        ["Annual Coverage Limit", f"PKR {quote['annual_limit']:,}"],
        ["Age Band", quote["age_band"]],
    ]
    plan_table = Table(plan_rows, colWidths=[70 * mm, 90 * mm])
    plan_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), slate),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#e7ebed")),
    ]))
    elements.append(plan_table)

    elements.append(Spacer(1, 16))
    premium_table = Table(
        [["Annual Premium", f"PKR {quote['premium']:,}"]],
        colWidths=[70 * mm, 90 * mm],
    )
    premium_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), light_bg),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("TEXTCOLOR", (0, 0), (0, 0), slate),
        ("TEXTCOLOR", (1, 0), (1, 0), green),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#d99aa4")),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
    ]))
    elements.append(premium_table)

    elements.append(Paragraph(
        f"Generated on {quote['generated_at']} &mdash; Personal Shield Premium Calculator. "
        "This is a computer-generated quotation for informational purposes and is subject to "
        "underwriting and policy terms.",
        footer_style,
    ))

    doc.build(elements)
    buffer.seek(0)

    filename = f"PersonalShield_Quotation_{quote['treatment_limit']}_{quote['dob']}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/export/excel", methods=["POST"])
def api_export_excel():
    data = request.get_json(silent=True) or {}
    dob_str = data.get("dob", "")
    treatment_limit = data.get("treatment_limit", "")

    try:
        quote = build_quote(dob_str, treatment_limit)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    rose_fill = PatternFill(start_color="9C5A68", end_color="9C5A68", fill_type="solid")
    light_fill = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
    green_font = Font(color="3B8F5C", bold=True, size=13)
    header_font = Font(color="FFFFFF", bold=True, size=14)
    label_font = Font(color="5C6B73", bold=True, size=10)
    value_font = Font(color="202A30", size=10)
    thin_border = Border(bottom=Side(style="thin", color="E7EBED"))

    ws.merge_cells("A1:B1")
    ws["A1"] = "Personal Shield — Premium Quotation"
    ws["A1"].font = header_font
    ws["A1"].fill = rose_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    rows = [
        ("Date of Birth", quote["dob"]),
        ("Current Age", quote["current_age"]),
        ("Age Next Birthday", quote["age_next_birthday"]),
        ("Treatment Limit", quote["treatment_limit"]),
        ("Annual Coverage Limit (PKR)", quote["annual_limit"]),
        ("Age Band", quote["age_band"]),
    ]

    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=value).font = value_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Annual Premium (PKR)").font = label_font
    ws.cell(row=r, column=2, value=quote["premium"]).font = green_font
    ws.cell(row=r, column=1).fill = light_fill
    ws.cell(row=r, column=2).fill = light_fill

    r += 2
    ws.cell(row=r, column=1, value=f"Generated on {quote['generated_at']}").font = Font(
        italic=True, color="5C6B73", size=8.5
    )

    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(2)].width = 26

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PersonalShield_Quotation_{quote['treatment_limit']}_{quote['dob']}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
